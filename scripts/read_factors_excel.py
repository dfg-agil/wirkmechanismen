import sys
from pathlib import Path
p = Path(__file__).resolve().parents[1] / 'scripts' / 'factors_measurability_proposals_RO_green.xlsx'
if not p.exists():
    print('MISSING_FILE', p)
    sys.exit(2)

# Try pandas first
try:
    import pandas as pd
    df = pd.read_excel(p)
    # normalize columns
    print('USING pandas')
    print(df.head(50).to_csv(index=False, sep=';'))
    sys.exit(0)
except Exception as e:
    err_pandas = str(e)

# Fallback to openpyxl
try:
    from openpyxl import load_workbook
    wb = load_workbook(filename=str(p), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    # print header
    def to_csv_line(vals):
        out = []
        for v in vals:
            if v is None:
                out.append('')
            else:
                s = str(v)
                if ';' in s or '\n' in s or '"' in s:
                    s = '"' + s.replace('"','""') + '"'
                out.append(s)
        return ';'.join(out)
    print('USING openpyxl')
    for i, r in enumerate(rows[:100]):
        print(to_csv_line(r))
    sys.exit(0)
except Exception as e:
    err_openpyxl = str(e)
# Final fallback: parse .xlsx as zip with XML
try:
    import zipfile
    import xml.etree.ElementTree as ET

    z = zipfile.ZipFile(str(p))
    # load shared strings if present
    shared = []
    if 'xl/sharedStrings.xml' in z.namelist():
        with z.open('xl/sharedStrings.xml') as fh:
            tree = ET.parse(fh)
            root = tree.getroot()
            for si in root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}si'):
                # concatenate text nodes
                texts = []
                for t in si.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t'):
                    texts.append(t.text or '')
                shared.append(''.join(texts))

    # Attempt to read first worksheet file
    sheet_name = None
    # prefer sheet1.xml
    candidates = [n for n in z.namelist() if n.startswith('xl/worksheets/sheet') and n.endswith('.xml')]
    if not candidates:
        raise RuntimeError('No sheet XML found in workbook')
    sheet_name = sorted(candidates)[0]
    with z.open(sheet_name) as fh:
        tree = ET.parse(fh)
        root = tree.getroot()
        ns = {'d': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
        rows = []
        for row in root.findall('.//d:sheetData/d:row', ns):
            cells = {}
            max_col = 0
            for c in row.findall('d:c', ns):
                ref = c.get('r')  # e.g., A1
                col = ''.join([ch for ch in ref if ch.isalpha()])
                # convert col to index
                idx = 0
                for ch in col:
                    idx = idx * 26 + (ord(ch.upper()) - ord('A') + 1)
                if idx > max_col:
                    max_col = idx
                cell_type = c.get('t')
                v = c.find('d:v', ns)
                value = ''
                if v is not None and v.text is not None:
                    if cell_type == 's':
                        try:
                            si = int(v.text)
                            value = shared[si] if si < len(shared) else ''
                        except Exception:
                            value = v.text
                    else:
                        value = v.text
                else:
                    # inline string?
                    is_node = c.find('d:is', ns)
                    if is_node is not None:
                        t = is_node.find('.//d:t', ns)
                        if t is not None and t.text:
                            value = t.text
                cells[idx - 1] = value
            # build row list
            row_list = [cells.get(i, '') for i in range(max_col)]
            rows.append(row_list)

    def csv_line(vals):
        out = []
        for v in vals:
            s = '' if v is None else str(v)
            if ';' in s or '"' in s or '\n' in s:
                s = '"' + s.replace('"', '""') + '"'
            out.append(s)
        return ';'.join(out)

    print('USING zip-xml-fallback; sheet=', sheet_name)
    for r in rows[:200]:
        print(csv_line(r))
    sys.exit(0)
except Exception as e:
    print('ERROR: could not read Excel')
    print('pandas error:', err_pandas)
    print('openpyxl error:', err_openpyxl)
    print('zip/xml error:', str(e))
    sys.exit(3)
