import argparse
import json
import re
from pathlib import Path
from typing import Dict, List, Set, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


KT_KEYWORDS = [
    r"koordin",
    r"okhuysen",
    r"bechky",
    r"dings[øo]yr",
    r"accountability",
    r"predictability",
    r"common understanding",
    r"boundary\s*spann",
    r"selbstkoordination",
]

MRT_KEYWORDS = [
    r"media\s*richness",
    r"\bmrt\b",
    r"daft",
    r"lengel",
    r"schmidt",
    r"ishii",
    r"medienreich",
    r"channel\s*expansion",
    r"carlson",
    r"zmud",
    r"equivocality",
    r"mehrdeutigkeit",
    r"medium-aufgabe-passung",
]


def find_elem_ids_in_markdown(path: Path) -> Set[str]:
    if not path.exists():
        return set()
    text = path.read_text(encoding="utf-8", errors="replace")
    return set(re.findall(r"elem-[A-Za-z0-9]+", text))


def text_blob_for_element(elem: dict) -> str:
    attrs = elem.get("attributes", {})
    tags = attrs.get("tags", [])
    if not isinstance(tags, list):
        tags = [str(tags)]
    parts = [
        elem.get("_id", ""),
        attrs.get("label", ""),
        attrs.get("description", ""),
        " ".join(str(t) for t in tags),
    ]
    return "\n".join(parts)


def text_blob_for_connection(conn: dict) -> str:
    attrs = conn.get("attributes", {})
    parts = [
        conn.get("_id", ""),
        attrs.get("label", ""),
        attrs.get("description", ""),
        attrs.get("connection type", ""),
    ]
    return "\n".join(parts)


def match_keywords(text: str, patterns: List[str]) -> Set[str]:
    lowered = text.casefold()
    hits = set()
    for pat in patterns:
        if re.search(pat, lowered, flags=re.IGNORECASE):
            hits.add(pat)
    return hits


def theory_from_text(text: str) -> Tuple[Set[str], Dict[str, List[str]]]:
    matches = {"KT": sorted(match_keywords(text, KT_KEYWORDS)), "MRT": sorted(match_keywords(text, MRT_KEYWORDS))}
    theories = set(k for k, v in matches.items() if v)
    return theories, matches


def theory_label(theories: Set[str]) -> str:
    if theories == {"KT", "MRT"}:
        return "KT+MRT"
    if theories == {"KT"}:
        return "KT"
    if theories == {"MRT"}:
        return "MRT"
    return ""


def autosize(ws, min_width: int = 10, max_width: int = 70):
    for col_cells in ws.columns:
        max_len = 0
        col = col_cells[0].column
        for c in col_cells:
            value = "" if c.value is None else str(c.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[get_column_letter(col)].width = max(min_width, min(max_width, max_len + 2))


def main():
    parser = argparse.ArgumentParser(description="Export KT/MRT-related factors and connections to Excel.")
    parser.add_argument(
        "--model",
        default="models/main_model/wirkmechanismen-main-model-blueprint.json",
        help="Path to main model blueprint JSON",
    )
    parser.add_argument(
        "--out",
        default="theory_subset_koordination_mrt.xlsx",
        help="Output .xlsx file",
    )
    args = parser.parse_args()

    repo = Path(__file__).resolve().parents[1]
    model_path = (repo / args.model).resolve()
    out_path = (repo / args.out).resolve()

    koord_doc_ids = find_elem_ids_in_markdown(repo / "KOORDINATIONSTHEORIE_COVERAGE_ANALYSIS.md")
    mrt_doc_ids = find_elem_ids_in_markdown(repo / "MEDIA_RICHNESS_THEORY_COVERAGE_ANALYSIS.md")

    data = json.loads(model_path.read_text(encoding="utf-8"))
    elements = data.get("elements", [])
    connections = data.get("connections", [])

    element_by_id = {e.get("_id"): e for e in elements}

    selected_factors: List[dict] = []
    selected_ids: Set[str] = set()
    element_theory: Dict[str, Set[str]] = {}

    for elem in elements:
        elem_id = elem.get("_id", "")
        attrs = elem.get("attributes", {})
        if attrs.get("element type") != "Einflussfaktoren":
            continue

        text = text_blob_for_element(elem)
        theories, matches = theory_from_text(text)

        if elem_id in koord_doc_ids:
            theories.add("KT")
            matches.setdefault("KT", []).append("coverage_doc_id")
        if elem_id in mrt_doc_ids:
            theories.add("MRT")
            matches.setdefault("MRT", []).append("coverage_doc_id")

        if not theories:
            continue

        selected_ids.add(elem_id)
        element_theory[elem_id] = theories

        tags = attrs.get("tags", [])
        if not isinstance(tags, list):
            tags = [str(tags)]

        selected_factors.append(
            {
                "element_id": elem_id,
                "label": attrs.get("label", ""),
                "element_type": attrs.get("element type", ""),
                "theory_basis": theory_label(theories),
                "kt_match": ", ".join(sorted(set(matches.get("KT", [])))),
                "mrt_match": ", ".join(sorted(set(matches.get("MRT", [])))),
                "in_koordination_coverage_doc": "yes" if elem_id in koord_doc_ids else "no",
                "in_mrt_coverage_doc": "yes" if elem_id in mrt_doc_ids else "no",
                "tags": " | ".join(str(t) for t in tags),
                "measurability": attrs.get("measurability"),
                "influenceability": attrs.get("influenceability"),
                "degree": attrs.get("degree"),
                "indegree": attrs.get("indegree"),
                "outdegree": attrs.get("outdegree"),
                "description": attrs.get("description", ""),
            }
        )

    selected_factors.sort(key=lambda r: (r["theory_basis"], r["label"]))

    selected_connections: List[dict] = []
    for conn in connections:
        from_id = conn.get("from")
        to_id = conn.get("to")

        if from_id not in selected_ids or to_id not in selected_ids:
            continue

        from_elem = element_by_id.get(from_id, {})
        to_elem = element_by_id.get(to_id, {})
        from_label = from_elem.get("attributes", {}).get("label", "")
        to_label = to_elem.get("attributes", {}).get("label", "")

        conn_text = text_blob_for_connection(conn)
        conn_theories, conn_matches = theory_from_text(conn_text)

        conn_theories = conn_theories.union(element_theory.get(from_id, set())).union(element_theory.get(to_id, set()))

        attrs = conn.get("attributes", {})
        selected_connections.append(
            {
                "connection_id": conn.get("_id", ""),
                "from_id": from_id,
                "from_label": from_label,
                "to_id": to_id,
                "to_label": to_label,
                "theory_basis": theory_label(conn_theories),
                "connection_type": attrs.get("connection type", ""),
                "source_label": attrs.get("label", ""),
                "kt_match": ", ".join(sorted(set(conn_matches.get("KT", [])))),
                "mrt_match": ", ".join(sorted(set(conn_matches.get("MRT", [])))),
                "description": attrs.get("description", ""),
            }
        )

    selected_connections.sort(key=lambda r: (r["theory_basis"], r["from_label"], r["to_label"]))

    wb = Workbook()

    ws_f = wb.active
    ws_f.title = "Einflussfaktoren"
    factor_headers = [
        "element_id",
        "label",
        "element_type",
        "theory_basis",
        "kt_match",
        "mrt_match",
        "in_koordination_coverage_doc",
        "in_mrt_coverage_doc",
        "tags",
        "measurability",
        "influenceability",
        "degree",
        "indegree",
        "outdegree",
        "description",
    ]
    ws_f.append(factor_headers)
    for h in ws_f[1]:
        h.font = Font(bold=True)
    for row in selected_factors:
        ws_f.append([row.get(k, "") for k in factor_headers])
    ws_f.freeze_panes = "A2"
    ws_f.auto_filter.ref = ws_f.dimensions
    autosize(ws_f)

    ws_c = wb.create_sheet("Verbindungen")
    conn_headers = [
        "connection_id",
        "from_id",
        "from_label",
        "to_id",
        "to_label",
        "theory_basis",
        "connection_type",
        "source_label",
        "kt_match",
        "mrt_match",
        "description",
    ]
    ws_c.append(conn_headers)
    for h in ws_c[1]:
        h.font = Font(bold=True)
    for row in selected_connections:
        ws_c.append([row.get(k, "") for k in conn_headers])
    ws_c.freeze_panes = "A2"
    ws_c.auto_filter.ref = ws_c.dimensions
    autosize(ws_c)

    ws_m = wb.create_sheet("Meta")
    ws_m.append(["Analysebasis", str(model_path.relative_to(repo))])
    ws_m.append(["Faktoren gesamt (Einflussfaktoren, theoriebezogen)", len(selected_factors)])
    ws_m.append(["Verbindungen zwischen theoriebezogenen Einflussfaktoren", len(selected_connections)])
    ws_m.append(["Koordinationstheorie-Orientierung", "Okhuysen & Bechky (2009); Dingsøyr et al. (2018)"])
    ws_m.append(["Media Richness Orientierung", "Daft & Lengel (1986); Schmidt et al. (2017); Ishii et al. (2019)"])
    ws_m.append(["Zusatzquelle im Repo", "KOORDINATIONSTHEORIE_COVERAGE_ANALYSIS.md, MEDIA_RICHNESS_THEORY_COVERAGE_ANALYSIS.md"])
    for h in ws_m[1]:
        h.font = Font(bold=True)
    autosize(ws_m)

    wb.save(out_path)

    print(f"WROTE: {out_path}")
    print(f"FACTORS: {len(selected_factors)}")
    print(f"CONNECTIONS: {len(selected_connections)}")


if __name__ == "__main__":
    main()
